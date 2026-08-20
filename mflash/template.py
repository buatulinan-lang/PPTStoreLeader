# -*- coding: utf-8 -*-
"""Template Excel untuk slide Struktur Organisasi: pembuat file & pembacanya."""
from __future__ import annotations
import io
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Arial"
NAVY = "1F3864"
NAVY2 = "2E5394"
CARD = "F6F8FC"
GARIS = "D5DDEB"

JABATAN_BAKU = (
    ["Ustadz Pembina Cabang", "Store Leader",
     "Supervisor Service", "Supervisor Aksesoris", "Supervisor Pengadaan",
     "Supervisor Penyewaan", "Supervisor Maintenance", "Supervisor ISP"]
    + ["Admin"] * 3
    + ["Sales"] * 3
    + ["Teknisi"] * 7
    + ["Sales Corporate"]
)

PILIHAN_JABATAN = ["Ustadz Pembina Cabang", "Store Leader",
                   "Supervisor Service", "Supervisor Aksesoris", "Supervisor Pengadaan",
                   "Supervisor Penyewaan", "Supervisor Maintenance", "Supervisor ISP",
                   "Admin", "Sales", "Teknisi", "Sales Counter", "Kasir", "Sales Corporate"]

PETUNJUK = [
    ("Cara pakai", ""),
    ("1", "Isi kolom NAMA LENGKAP pada sheet STRUKTUR. Kolom JABATAN sudah tersedia daftarnya."),
    ("2", "Baris yang namanya dikosongkan tetap muncul di slide sebagai kotak tanpa nama. "
          "Hapus barisnya bila posisi itu memang tidak ada di cabang Anda."),
    ("3", "Baris sudah disiapkan: 3 Admin, 3 Sales, dan 7 Teknisi. Boleh ditambah atau dikurangi "
          "sesuai jumlah SDM di cabang — semua nama akan muncul di slide."),
    ("4", "Simpan file, lalu unggah di aplikasi pada tab Slide manual → Struktur organisasi."),
    ("", ""),
    ("Aturan penempatan otomatis", ""),
    ("Ustadz / Pembina", "Puncak bagan"),
    ("Store Leader", "Tepat di bawah Ustadz Pembina Cabang"),
    ("Supervisor ...", "Berjajar sesuai urutan: Service, Aksesoris, Pengadaan, Penyewaan, "
                       "Maintenance, ISP"),
    ("Sales Corporate", "Di bawah Supervisor Pengadaan, Penyewaan, Maintenance, dan ISP"),
    ("Jabatan lain", "Di bawah Supervisor Service, dikelompokkan per jabatan — satu kartu berisi "
                     "daftar nama (Admin, Sales, Teknisi, dst.)"),
    ("", ""),
    ("Contoh pengisian", ""),
    ("Ust. Abdullah Hakim", "Ustadz Pembina Cabang"),
    ("Ahmad Fauzi", "Store Leader"),
    ("Budi Santoso", "Supervisor Service"),
    ("Hana Safitri", "Admin"),
]


def buat_template_struktur() -> bytes:
    """Hasilkan file Excel siap isi untuk slide Struktur Organisasi."""
    wb = Workbook()

    ws = wb.active
    ws.title = "STRUKTUR"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "TEMPLATE STRUKTUR ORGANISASI — M-FLASH"
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws["A2"] = ("Isi kolom NAMA LENGKAP (kolom biru muda). Kolom JABATAN memakai daftar pilihan. "
                "Baca sheet PETUNJUK bila ragu.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="6B7280")

    kepala = ["NAMA LENGKAP", "JABATAN"]
    for i, h in enumerate(kepala, start=1):
        sel = ws.cell(row=4, column=i, value=h)
        sel.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        sel.fill = PatternFill("solid", fgColor=NAVY)
        sel.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 22

    tipis = Side(style="thin", color=GARIS)
    tepi = Border(left=tipis, right=tipis, top=tipis, bottom=tipis)
    isian = PatternFill("solid", fgColor="EAF3FB")

    for baris, jab in enumerate(JABATAN_BAKU, start=5):
        a = ws.cell(row=baris, column=1, value=None)
        b = ws.cell(row=baris, column=2, value=jab)
        a.fill = isian
        for sel in (a, b):
            sel.font = Font(name=FONT, size=10)
            sel.border = tepi
            sel.alignment = Alignment(vertical="center")
        ws.row_dimensions[baris].height = 19

    baris_akhir = 4 + len(JABATAN_BAKU)
    # baris kosong tambahan supaya mudah menambah orang
    for baris in range(baris_akhir + 1, baris_akhir + 11):
        a = ws.cell(row=baris, column=1)
        b = ws.cell(row=baris, column=2)
        a.fill = isian
        for sel in (a, b):
            sel.font = Font(name=FONT, size=10)
            sel.border = tepi
        ws.row_dimensions[baris].height = 19

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 30

    # daftar pilihan jabatan (sheet tersembunyi)
    wl = wb.create_sheet("DAFTAR_JABATAN")
    for i, j in enumerate(PILIHAN_JABATAN, start=1):
        wl.cell(row=i, column=1, value=j).font = Font(name=FONT, size=10)
    wl.column_dimensions["A"].width = 30
    wl.sheet_state = "hidden"

    dv = DataValidation(type="list",
                        formula1=f"=DAFTAR_JABATAN!$A$1:$A${len(PILIHAN_JABATAN)}",
                        allow_blank=True, showDropDown=False)
    dv.prompt = "Pilih jabatan dari daftar, atau ketik jabatan lain."
    dv.promptTitle = "Jabatan"
    ws.add_data_validation(dv)
    dv.add(f"B5:B{baris_akhir + 10}")

    # sheet petunjuk
    wp = wb.create_sheet("PETUNJUK")
    wp.sheet_view.showGridLines = False
    wp["A1"] = "PETUNJUK PENGISIAN"
    wp["A1"].font = Font(name=FONT, size=13, bold=True, color=NAVY)
    for i, (kiri, kanan) in enumerate(PETUNJUK, start=3):
        a = wp.cell(row=i, column=1, value=kiri)
        b = wp.cell(row=i, column=2, value=kanan)
        tebal = kanan == "" and kiri != ""
        a.font = Font(name=FONT, size=10, bold=tebal, color=NAVY if tebal else "20242E")
        b.font = Font(name=FONT, size=10, color="20242E")
        b.alignment = Alignment(wrap_text=True, vertical="top")
    wp.column_dimensions["A"].width = 26
    wp.column_dimensions["B"].width = 78

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def baca_struktur(file) -> list:
    """Baca template yang sudah diisi menjadi daftar {nama, jabatan}."""
    nama_file = str(getattr(file, "name", file)).lower()
    if nama_file.endswith(".csv"):
        df = pd.read_csv(file)
    else:
        wb = load_workbook(file, data_only=True)
        ws = wb["STRUKTUR"] if "STRUKTUR" in wb.sheetnames else wb[wb.sheetnames[0]]
        baris = list(ws.values)
        kepala_idx = None
        for i, r in enumerate(baris):
            nilai = [str(x).strip().upper() if x is not None else "" for x in r]
            if "JABATAN" in nilai and any("NAMA" in v for v in nilai):
                kepala_idx = i
                break
        if kepala_idx is None:
            return []
        kolom = [str(x).strip().upper() if x is not None else "" for x in baris[kepala_idx]]
        df = pd.DataFrame(baris[kepala_idx + 1:], columns=kolom)

    df.columns = [str(c).strip().upper() for c in df.columns]
    kol_nama = next((c for c in df.columns if "NAMA" in c), None)
    kol_jab = next((c for c in df.columns if "JABATAN" in c), None)
    if not kol_jab:
        return []
    out = []
    for _, r in df.iterrows():
        nama = "" if pd.isna(r.get(kol_nama)) else str(r.get(kol_nama)).strip()
        jab = "" if pd.isna(r.get(kol_jab)) else str(r.get(kol_jab)).strip()
        if not jab or jab.upper() in ("NAN", "NONE"):
            continue
        if nama.upper().startswith("CONTOH"):
            continue
        out.append({"nama": nama, "jabatan": jab})
    return out
